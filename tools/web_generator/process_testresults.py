import os
import sys
import openpyxl
import pickle

class result_data:
    def __init__(self):
        self.data = []
        
        self.data = [{'Function Group': 'Test Group', 'Pass': 1, 'Fail': 1, 'Not Test': 1}]
        self.data_detail = [{'Function Group': 'Test Group', 'Pass': 1, 'Fail': 1, 'Not Test': 1, 'None': 1, 'Remark': 'None'}]
        
    def add_data_detail(self, input={'Function Group': 'Test Group', 'Pass': 1, 'Remark': 'None'}):
        
        found_result_group = False
        for idx, onegroup in enumerate(self.data_detail):
            if onegroup['Function Group'] ==  input['Function Group'] and onegroup['Remark'] ==  input['Remark']:
                found_result_group = True
                if 'Pass' in input:
                    onegroup['Pass'] += 1
                elif 'Fail' in input:
                    onegroup['Fail'] += 1
                elif 'Not Test' in input:
                    onegroup['Not Test'] += 1
                elif 'None' in input:
                    onegroup['None'] += 1
                else:
                    onegroup['None'] += 1
                self.data_detail[idx] = onegroup
                break
        if found_result_group == False:
            temp_dict = input
            if 'Pass' not in input:
                temp_dict['Pass'] = 0
            if 'Fail' not in input:
                temp_dict['Fail'] = 0
            if 'Not Test' not in input:
                temp_dict['Not Test'] = 0
            if 'None' not in input:
                temp_dict['None'] = 0
            self.data_detail.append(temp_dict)
            
    def add_data(self, input={'Function Group': 'Test Group', 'Pass': 1}):
        found_function_group = False
        for idx, onegroup in enumerate(self.data):
            if onegroup['Function Group'] ==  input['Function Group']:
                found_function_group = True
                if 'Pass' in input:
                    onegroup['Pass'] += 1
                elif 'Fail' in input:
                    onegroup['Fail'] += 1
                elif 'Not Test' in input:
                    onegroup['Not Test'] += 1
                else:
                    onegroup['Not Test'] += 1
                self.data[idx] = onegroup
                print onegroup
                break
        if found_function_group == False:
            temp_dict = input
            if 'None' in temp_dict:
                del temp_dict['None']
                temp_dict['Not Test'] = 1
            if 'Pass' not in input:
                temp_dict['Pass'] = 0
            if 'Fail' not in input:
                temp_dict['Fail'] = 0
            if 'Not Test' not in input:
                temp_dict['Not Test'] = 0
            self.data.append(temp_dict)

class testresults:
    def __init__(self, file_path):
        self.xl_file = file_path

        self.data = result_data()
        
        self.workbook = openpyxl.load_workbook(self.xl_file)

        self.worksheet = self.workbook.get_sheet_by_name('Function Test')
        
        print self.workbook
        print self.worksheet
        print self.worksheet.title  
        print self.worksheet.max_row
        print self.worksheet.max_column
        #workbook = xlrd.open_workbook(self.xl_file)
        #sheet_names= workbook.sheet_names()
        #print sheet_names
        #for sheet_name in sheet_names:
        #    if u'Function Test' == sheet_name:
        #        self.functiontest_sheet = workbook.sheet_by_name(sheet_name)
        #    print sheet_name


    def get_all_importent_data(self):
        max_row = self.worksheet.max_row
        max_column = self.worksheet.max_column
        
        result_data_list = []
        for column in range(max_column):
            if 'Function Group' == self.worksheet.cell(row = 1, column = column+1).value:
                print '####This is function Group'
                function_group_list = []
                for row in range(1, max_row):
                    cell_value = self.worksheet.cell(row = row+1, column = column+1).value
                    cell_value =str(cell_value)
                    print cell_value
                    function_group_list.append(cell_value)
                result_data_list.append(function_group_list)
            if 'Test Result' == self.worksheet.cell(row = 1, column = column+1).value:
                print '####This is Test Result'
                test_result_list = []
                for row in range(1, max_row):
                    cell_value = self.worksheet.cell(row = row+1, column = column+1).value
                    cell_value =str(cell_value)
                    print cell_value
                    test_result_list.append(cell_value)
                result_data_list.append(test_result_list)
            if 'Remark' == self.worksheet.cell(row = 1, column = column+1).value:
                print '####This is Remark'
                remark_list = []
                for row in range(1, max_row):
                    cell_value = self.worksheet.cell(row = row+1, column = column+1).value
                    cell_value =str(cell_value)
                    print cell_value
                    remark_list.append(cell_value)
                result_data_list.append(remark_list)
        
        print result_data_list
        for idx in range(len(result_data_list[0])):
            temp_dict = {'Function Group': result_data_list[0][idx], result_data_list[1][idx]: 1, 'Remark': result_data_list[2][idx]}
            self.data.add_data(temp_dict)
        for idx in range(len(result_data_list[0])):
            temp_dict = {'Function Group': result_data_list[0][idx], result_data_list[1][idx]: 1, 'Remark': result_data_list[2][idx]}
            self.data.add_data_detail(temp_dict)
        print '-------------------------------------'
        for item in self.data.data:
            print item
        print '-------------------------------------'
        for item in self.data.data_detail:
            print item
        print '-------------------------------------'
        return self.data.data, self.data.data_detail

    def display_all_content_in_sheet(self):
        max_row = self.worksheet.max_row
        max_column = self.worksheet.max_column
        for row in range(max_row):
            for column in range(max_column):
                print self.worksheet.cell(row = row+1, column = column+1).value 

if __name__ == "__main__":
    print "hello world!"
    #new_xl = testresults(r'\\10.170.2.9\file_share_vol\D03_EE&Info&Con&Cloud\3.Electrical & Electronics\3.5 EEV\3.5.5 Labcar\04_TestCase\FRS5.0\FT\Test Case\M31T FRS5.0 System Test Report_20170510.xlsx')
    new_xl = testresults(r'D:\report.xlsx')
    print new_xl.get_all_importent_data()